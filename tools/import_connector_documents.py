#!/usr/bin/env python3
"""Create a reviewable connector-import staging JSON from common document types."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from pathlib import Path


CONTACT_RE = re.compile(r"^\s*(?:pin|contacto|contact)?\s*([A-Z]?\d{1,3}|[A-Z]{1,4}[+-]?)\s*[:;|,\-]\s*(.{1,180})$", re.I)
ID_KEYS = ("connector_id", "conector_id", "connector", "conector", "id_conector")
CONTACT_KEYS = ("contact_id", "pin", "contacto", "contact", "terminal")
SIGNAL_KEYS = ("signal", "senal", "señal", "function", "funcion", "función")
NAME_KEYS = ("canonical_name", "name", "nombre", "conector")


def first(row: dict, keys: tuple[str, ...]) -> str:
    lowered = {str(key).strip().lower(): value for key, value in row.items()}
    for key in keys:
        value = lowered.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def rows_to_candidates(rows: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    for index, row in enumerate(rows, start=1):
        connector_id = first(row, ID_KEYS) or f"pending-connector-{index}"
        candidate = grouped.setdefault(connector_id, {
            "proposed_id": connector_id,
            "canonical_name": first(row, NAME_KEYS) or connector_id,
            "contacts": [],
            "review_status": "pending_review",
        })
        contact_id = first(row, CONTACT_KEYS)
        signal = first(row, SIGNAL_KEYS)
        if contact_id or signal:
            candidate["contacts"].append({"id": contact_id or str(len(candidate["contacts"]) + 1), "signal": signal or "PENDIENTE", "description": "Importado; requiere contraste."})
    return list(grouped.values())


def parse_delimited(path: Path) -> tuple[list[dict], dict]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    rows = list(csv.DictReader(text.splitlines(), dialect=dialect))
    return rows_to_candidates(rows), {"rows": len(rows), "delimiter": dialect.delimiter}


def parse_json(path: Path) -> tuple[list[dict], dict]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    records = payload if isinstance(payload, list) else payload.get("records", []) if isinstance(payload, dict) else []
    if records and all(isinstance(item, dict) and "contacts" in item for item in records):
        candidates = records
    else:
        candidates = rows_to_candidates([item for item in records if isinstance(item, dict)])
    return candidates, {"root_type": type(payload).__name__, "source_records": len(records)}


def parse_xlsx(path: Path) -> tuple[list[dict], dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl no está disponible; deja el lote como needs_extractor") from error
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    sheets = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        headers = [str(value or "").strip() for value in values[0]]
        usable = [index for index, header in enumerate(headers) if header]
        if not usable:
            continue
        sheets.append(sheet.title)
        for values_row in values[1:]:
            row = {headers[index]: values_row[index] for index in usable if index < len(values_row)}
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
    return rows_to_candidates(rows), {"sheets": sheets, "rows": len(rows)}


def text_candidates(text_by_page: list[tuple[int, str]]) -> tuple[list[dict], dict]:
    contacts = []
    for page, text in text_by_page:
        for line in text.splitlines():
            match = CONTACT_RE.match(line)
            if match:
                contacts.append({"id": match.group(1).upper(), "signal": match.group(2).strip(), "description": f"Candidato extraído de la página {page}; requiere revisión."})
    candidates = []
    if contacts:
        candidates.append({"proposed_id": "pending-from-text", "canonical_name": "Conector pendiente de identificar", "contacts": contacts[:200], "review_status": "pending_review"})
    return candidates, {"pages_or_sections": len(text_by_page), "contact_lines_detected": len(contacts)}


def parse_pdf(path: Path) -> tuple[list[dict], dict]:
    try:
        from pypdf import PdfReader
    except ImportError as error:
        raise RuntimeError("pypdf no está disponible; deja el lote como needs_extractor") from error
    reader = PdfReader(str(path))
    pages = [(index + 1, page.extract_text() or "") for index, page in enumerate(reader.pages)]
    candidates, meta = text_candidates(pages)
    meta["pages"] = len(reader.pages)
    return candidates, meta


def parse_text(path: Path) -> tuple[list[dict], dict]:
    return text_candidates([(1, path.read_text(encoding="utf-8-sig", errors="replace"))])


def build_staging(path: Path) -> dict:
    suffix = path.suffix.lower()
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    status = "extracted"
    warning = None
    if suffix == ".json":
        candidates, extraction = parse_json(path)
    elif suffix in {".csv", ".tsv"}:
        candidates, extraction = parse_delimited(path)
    elif suffix == ".xlsx":
        candidates, extraction = parse_xlsx(path)
    elif suffix == ".pdf":
        candidates, extraction = parse_pdf(path)
    elif suffix == ".txt":
        candidates, extraction = parse_text(path)
    elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        candidates, extraction, status = [], {"vision_analysis_required": True}, "needs_extractor"
        warning = "La imagen se conserva, pero una IA con visión debe extraer forma, orientación y contactos."
    else:
        raise ValueError(f"Formato no admitido: {suffix}")
    if status == "extracted" and not candidates:
        status = "needs_extractor"
        warning = warning or "No se detectaron filas de contactos con suficiente estructura."
    return {
        "schema_version": "1.0",
        "kind": "connector_import_staging",
        "source": {"filename": path.name, "sha256": digest, "size": path.stat().st_size, "extension": suffix.lstrip(".")},
        "import_status": status,
        "extraction": extraction,
        "candidates": candidates,
        "warning": warning,
        "merge_policy": "Nunca sobrescribir el catálogo: comparar ID, variante, orientación, contactos y fuentes; toda diferencia pasa a revisión humana.",
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    if not args.input.is_file():
        raise SystemExit(f"No existe: {args.input}")
    output = args.output or args.input.with_suffix(args.input.suffix + ".staging.json")
    output.write_text(json.dumps(build_staging(args.input), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
